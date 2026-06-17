"""Build the uk_baseline data catalogue from the database's own metadata.

Input/output CRS: n/a (reads metadata only). Data layer: catalogue generation.

Single source of truth is the live PostGIS database: every layer's
``COMMENT ON TABLE`` and ``COMMENT ON COLUMN`` text. This script reads that
metadata (read-only, no writes) and emits, side by side:

  * ``catalogue/uk_baseline_data_dictionary.xlsx`` — one workbook, a Layers
    sheet (one row per layer) and a Columns sheet (one row per column).
  * ``catalogue/uk_baseline_catalogue.html`` — one self-contained, offline
    HTML page with a live search box; nothing to install, just open it.
  * ``catalogue/docs/`` + ``catalogue/mkdocs.yml`` — MkDocs Material source
    for the browsable GitHub Pages site (one page per layer).
  * ``catalogue/.catalogue_fingerprint.json`` — a hash of all metadata, so a
    later hook can tell whether the published catalogue has drifted from the DB.

Nothing about the database connection (host, user, credentials, internal file
paths) is ever written into a catalogue output — the published artefacts carry
data-dictionary content only.

British English throughout. Re-run any time the DB metadata changes.
"""
from __future__ import annotations

import hashlib
import html
import json
import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import psycopg
from dotenv import dotenv_values
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from psycopg.rows import dict_row

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

# --- configuration ---------------------------------------------------------
# Path to the .env holding PG_* connection settings. Override with the
# UK_BASELINE_ENV environment variable so the script is portable across machines.
ENV = Path(os.environ.get(
    "UK_BASELINE_ENV",
    r"C:\Users\po.nienchen\OneDrive - priorpartners.com"
    r"\Documents\05_Code_Scripts\data_import\Natural_England\.env"))
SCHEMA = "uk_baseline"
ROOT = Path(__file__).resolve().parents[2]          # the uk_baseline working dir
OUT = ROOT / "catalogue"                             # MkDocs site source (local)
DOCS = OUT / "docs"
# Colleague-facing single-file deliverables live on the practice share.
SHARE = Path(r"P:\0_Practice\10_Data management resources\04_Spreadsheets")
# Static preview snapshots (built by build_maps.py) live under docs/maps/ as
# <table>.png and deploy with the MkDocs site. Interactive / full-extent viewing
# is now the Dashboard's job (Stream 1); each layer links out to it.
MAPS_DIR = DOCS / "maps"
MAP_SITE_BASE = "https://prior-partners.github.io/PP_uk_baseline/maps/"
# The Dashboard runs per-user locally (BYO-local model); the base URL is the same
# localhost for every colleague. Deep-link keys on the bare uk_baseline table name.
DASHBOARD_BASE = "http://localhost:7800/"


def dashboard_link(table: str) -> str:
    """Deep-link that opens this layer in the colleague's local Dashboard."""
    return f"{DASHBOARD_BASE}?layer=uk_baseline.{table}"


# Shared styles file; its "_approved" list gates which themes' previews go live
# (themes are reviewed + deployed one at a time).
STYLES_JSON = Path(os.environ.get(
    "UK_BASELINE_STYLES",
    r"P:\0_Practice\10_Data management resources\XX_Working\uk_baseline_styles.json"))


def _approved_themes() -> set:
    try:
        return set(json.loads(STYLES_JSON.read_text(encoding="utf-8")).get("_approved", []))
    except Exception:
        return set()


APPROVED = _approved_themes()


def has_map(table: str) -> bool:
    """True if a preview snapshot exists AND its theme is approved for the site."""
    theme = table.split("_", 1)[0].upper()
    return theme in APPROVED and (MAPS_DIR / f"{table}.png").exists()


# Layer short names — a concise human label shown between the full title and the
# table name — come from the PP data register's ``data_short_name`` column (the
# firm's curated short-name field). The catalogue is otherwise DB-driven; this is
# the one external lookup, and a blank/missing value simply renders no line.
REGISTER = Path(os.environ.get(
    "UK_BASELINE_REGISTER",
    r"P:\0_Practice\10_Data management resources\04_Spreadsheets\DI"
    r"\uk_baseline_data_register.xlsx"))


def read_short_names() -> dict[str, str]:
    """Map ``table_name`` -> short name (first letter capitalised) from the register."""
    out: dict[str, str] = {}
    try:
        wb = load_workbook(REGISTER, read_only=True, data_only=True)
        ws = wb["uk_baseline_current"]
        rows = ws.iter_rows(values_only=True)
        hdr = list(next(rows))
        ti, si = hdr.index("table_name"), hdr.index("data_short_name")
        for row in rows:
            table, short = row[ti], row[si]
            if table and short and str(short).strip():
                s = str(short).strip()
                out[str(table)] = s[0].upper() + s[1:]
        wb.close()
    except Exception as exc:  # noqa: BLE001 - the register is an optional adornment
        log.warning("Short names unavailable (%s): %s", REGISTER.name, exc)
    return out


# Theme prefix -> full label. From the firm's 12-theme taxonomy (P+P brand /
# data-management standard), so colleagues unfamiliar with the codes can read
# the catalogue. HOU is in the taxonomy but has no layers loaded yet; it is
# only rendered when a HOU_* table exists.
THEME_LABELS: dict[str, str] = {
    "ADM": "Administrative & Planning Boundaries",
    "BLT": "Built Environment & Land Use",
    "COM": "Community & Social Infrastructure",
    "DEM": "Demographics & Socio-Economic",
    "ECN": "Economy & Employment",
    "EDU": "Education",
    "ENV": "Environment & Natural Assets",
    "HER": "Heritage & Historical Assets",
    "HOU": "Housing & Living Conditions",
    "HTH": "Health & Wellbeing",
    "MOB": "Transport & Mobility",
    "UTL": "Infrastructure & Utilities",
}

# Typical contents per theme, quoted from the firm's 12-theme taxonomy table.
# Shown as a theme blurb so readers grasp the theme's remit at a glance.
THEME_DESC: dict[str, str] = {
    "ADM": "Local authority, wards, output areas, built-up areas, postcode boundaries, planning zones, workplace zones",
    "BLT": "Building footprints, land use classifications, planning applications, retail centres, points of interest",
    "COM": "Sports facilities, libraries, community centres, places of worship, cultural venues",
    "DEM": "Census data (age, ethnicity, religion), deprivation indices, income data, household, travel-to-work data",
    "ECN": "Employment by sector, business registers, economic output, unemployment rates",
    "EDU": "Schools database, Ofsted ratings, university locations",
    "ENV": "Protected sites, woodland, green spaces, flood zones, air quality, noise monitoring, radiation monitoring",
    "HER": "Listed buildings, conservation areas, historic parks, archaeological sites",
    "HOU": "House prices, housing tenure, property types, energy efficiency ratings, overcrowding metrics, rental costs",
    "HTH": "GP surgeries, hospitals, health inequalities, disease prevalence, obesity",
    "MOB": "Road networks, railway lines, public transport routes, cycling infrastructure, stations, paths, trails",
    "UTL": "Electricity networks, gas pipelines, water infrastructure, telecoms masts, waste facilities",
}

# Per-theme colour, drawn from the P+P brand palette (P+P_Brand Guide_v1.pdf).
# Core brand: orange #e9511d, dark plum #2b242c, cream #f1eedf, black #121212.
# The eight brand hues below cover most themes; the last four are 0.6-shaded
# brand hues so every theme has a distinct, on-brand colour.
THEME_COLOURS: dict[str, str] = {
    "ADM": "#e9511d",  # brand orange
    "BLT": "#920f67",  # brand magenta
    "COM": "#554596",  # purple (wheel A/B)            — per data owner
    "DEM": "#945f14",  # brown (wheel I/H)             — per data owner
    "ECN": "#dd1220",  # orange-red (wheel L/K)        — per data owner
    "EDU": "#6da2ed",  # light blue (wheel C, light)   — per data owner
    "ENV": "#219900",  # green (wheel F)               — per data owner
    "HER": "#2b242c",  # brand dark plum
    "HOU": "#8c3111",  # shaded orange (no layers yet)
    "HTH": "#005dff",  # brand blue (per data owner, from supplied swatch)
    "MOB": "#c8c8c4",  # light grey                    — per data owner
    "UTL": "#991955",  # shaded pink
}

# Brand neutrals for the page chrome.
BRAND_INK = "#2b242c"
BRAND_ACCENT = "#e9511d"
BRAND_BG = "#f1eedf"
BRAND_BLACK = "#121212"


def text_on(hex_colour: str) -> str:
    """Pick black or white text for legibility on a background colour."""
    r, g, b = (int(hex_colour[i:i + 2], 16) for i in (1, 3, 5))
    luminance = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return BRAND_BLACK if luminance > 150 else "#ffffff"


def first_sentence(text: str) -> str:
    """First sentence of a description, used as the human-readable title."""
    text = text.strip()
    if not text:
        return ""
    end = text.find(". ")
    return text[:end + 1].strip() if end != -1 else text

# Section headers used in the firm's COMMENT template, in display order.
SECTION_HEADERS = [
    "SOURCE", "DOCUMENTATION", "DEFINITIONS", "SCOPE", "CRS", "LICENCE",
    "DATA QUALITY CAVEATS", "ENRICHMENT", "UPDATE REQUIRED",
    "NOT IN THIS DATASET", "DERIVED FROM", "LOADED INTO uk_baseline",
]


def conninfo() -> str:
    env = dotenv_values(ENV)
    parts = [f"host={env['PG_HOST']}", f"port={env['PG_PORT']}",
             f"dbname={env['PG_DATABASE']}", f"user={env['PG_USER']}"]
    if env.get("PG_PASSWORD"):
        parts.append(f"password={env['PG_PASSWORD']}")
    return " ".join(parts)


# --- comment parsing -------------------------------------------------------
def _is_header(line: str) -> str | None:
    """Return the canonical header if ``line`` starts a known section, else None."""
    stripped = line.strip()
    for head in SECTION_HEADERS:
        if stripped == head or stripped.startswith(head + " "):
            return head
    return None


def parse_table_comment(comment: str | None) -> dict[str, Any]:
    """Split a sectioned COMMENT into ``description`` plus a section -> text map.

    The text of each section is preserved verbatim (bullets kept) so nothing is
    paraphrased; rendering decides presentation.
    """
    if not comment:
        return {"description": "", "sections": {}}
    lines = comment.splitlines()
    desc: list[str] = []
    sections: dict[str, list[str]] = {}
    current: str | None = None
    for line in lines:
        head = _is_header(line)
        if head is not None:
            current = head
            sections.setdefault(current, [])
            continue
        if current is None:
            if line.strip():
                desc.append(line.strip())
        else:
            sections[current].append(line.rstrip())
    # Trim leading/trailing blank lines inside each section, join to text.
    clean: dict[str, str] = {}
    for head, body in sections.items():
        while body and not body[0].strip():
            body.pop(0)
        while body and not body[-1].strip():
            body.pop()
        clean[head] = "\n".join(body)
    return {"description": " ".join(desc).strip(), "sections": clean}


def parse_bullets(text: str) -> list[str]:
    """Section body -> list of complete bullets.

    A bullet in the firm's template can wrap across several indented lines; the
    continuation lines are folded back into the one bullet so nothing is split
    or dropped. A section with no ``- `` markers is folded into a single item.
    """
    bullets: list[str] = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("- "):
            bullets.append(s[2:].strip())
        elif bullets:
            bullets[-1] = f"{bullets[-1]} {s}"
        else:
            bullets.append(s)
    return bullets


def first_bullet(text: str) -> str:
    """First complete bullet of a section body; '' if none."""
    b = parse_bullets(text)
    return b[0] if b else ""


def first_url(text: str) -> str:
    """First http(s) URL in a section body (the datasource link); '' if none."""
    m = re.search(r"https?://[^\s)<]+", text or "")
    return m.group(0) if m else ""


# --- database read ---------------------------------------------------------
def read_metadata() -> list[dict[str, Any]]:
    """Read every base table in the schema with its comments and columns."""
    records: list[dict[str, Any]] = []
    with psycopg.connect(conninfo(), row_factory=dict_row) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT c.oid, c.relname AS table_name,
                       upper(split_part(c.relname, '_', 1)) AS theme,
                       obj_description(c.oid) AS table_comment
                FROM pg_class c
                JOIN pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = %s AND c.relkind = 'r'
                ORDER BY theme, c.relname
                """,
                (SCHEMA,),
            )
            tables = cur.fetchall()
            for t in tables:
                cur.execute(
                    """
                    SELECT a.attname AS name,
                           format_type(a.atttypid, a.atttypmod) AS type,
                           col_description(a.attrelid, a.attnum) AS comment
                    FROM pg_attribute a
                    WHERE a.attrelid = %s AND a.attnum > 0 AND NOT a.attisdropped
                    ORDER BY a.attnum
                    """,
                    (t["oid"],),
                )
                cols = cur.fetchall()
                parsed = parse_table_comment(t["table_comment"])
                records.append({
                    "theme": t["theme"],
                    "table": t["table_name"],
                    "raw_comment": t["table_comment"] or "",
                    "description": parsed["description"],
                    "sections": parsed["sections"],
                    "columns": cols,
                })
    return records


# --- fingerprint -----------------------------------------------------------
def fingerprint(records: list[dict[str, Any]]) -> str:
    """Stable SHA-256 over all table + column comment text."""
    canon = [
        {
            "table": r["table"],
            "comment": r["raw_comment"],
            "columns": [(c["name"], c["comment"] or "") for c in r["columns"]],
        }
        for r in sorted(records, key=lambda x: x["table"])
    ]
    blob = json.dumps(canon, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


# --- renderers -------------------------------------------------------------
def render_xlsx(records: list[dict[str, Any]], built: str, path: Path) -> None:
    head_fill = PatternFill("solid", fgColor="1F4E5F")
    head_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    bold = Font(bold=True)

    # Preserve the existing workbook (the "About & how to use" sheet is now
    # owned and hand-edited by the data owner — never overwrite it). Only the
    # generated Layers and Columns sheets are rebuilt. On first run (no file),
    # the About sheet is created once from the default template below.
    first_run = not path.exists()
    if first_run:
        wb = Workbook()
        wb.remove(wb.active)
    else:
        wb = load_workbook(path)
        for name in ("Layers", "Columns"):
            if name in wb.sheetnames:
                del wb[name]

    ws = wb.create_sheet("Layers")
    layer_cols = ["Theme", "Layer (table name)", "Description", "Source",
                  "Datasource link", "Definitions", "Scope", "CRS", "Licence",
                  "Update required", "No. of columns"]
    ws.append(layer_cols)
    for r in records:
        s = r["sections"]
        ws.append([
            r["theme"], r["table"], r["description"],
            first_bullet(s.get("SOURCE", "")),
            first_url(s.get("DOCUMENTATION", "")),
            first_bullet(s.get("DEFINITIONS", "")),
            s.get("SCOPE", "").replace("- ", "").strip(),
            first_bullet(s.get("CRS", "")),
            first_bullet(s.get("LICENCE", "")),
            first_bullet(s.get("UPDATE REQUIRED", "")),
            len(r["columns"]),
        ])

    wc = wb.create_sheet("Columns")
    col_cols = ["Theme", "Layer (table name)", "Column", "Type", "Description / unit"]
    wc.append(col_cols)
    for r in records:
        for c in r["columns"]:
            wc.append([r["theme"], r["table"], c["name"], c["type"],
                       c["comment"] or ""])

    if first_run:
        _build_about_sheet(wb, records, built, head_fill, head_font, bold)

    for sheet, widths in ((ws, [8, 46, 70, 40, 52, 50, 50, 22, 40, 40, 14]),
                          (wc, [8, 46, 26, 26, 80])):
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        for cell in sheet[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = wrap
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("Wrote %s (About sheet %s)", path.name,
             "created" if first_run else "preserved")


def _build_about_sheet(wb, records, built, head_fill, head_font, bold) -> None:
    """First-run-only default 'About & how to use' sheet (owner edits it after)."""
    info = wb.create_sheet("About & how to use", 0)
    n_themes = len({r["theme"] for r in records})
    head = Font(bold=True, color="1b2733", size=12)
    title = Font(bold=True, size=16, color="2b242c")
    rows: list[tuple[str, Font | None]] = [
        ("uk_baseline data catalogue", title),
        (f"{len(records)} layers across {n_themes} themes  ·  generated from the live database on {built}", None),
        ("", None),
        ("What this workbook is", head),
        ("A guide to every spatial layer held in the uk_baseline database. It is generated", None),
        ("automatically from each table's own documentation, so it always matches the live database —", None),
        ("the database is the single source of truth. Do not hand-edit this file; it is overwritten on each rebuild.", None),
        ("", None),
        ("The two other sheets", head),
        ("1.  \"Layers\"  — one row per layer (database table). Use this to find out WHAT layers exist.", bold),
        ("        Theme            two/three-letter theme code (see legend below).", None),
        ("        Layer            the database table name.", None),
        ("        Description      one-line summary of the layer.", None),
        ("        Source           the publishing organisation.", None),
        ("        Datasource link  the documentation / source URL for the dataset.", None),
        ("        Definitions      the publisher's own definition of the core concept.", None),
        ("        Scope            geographic coverage, filter, and row count.", None),
        ("        CRS              coordinate reference system (EPSG code).", None),
        ("        Licence          usage licence.", None),
        ("        Update required  whether a newer release is awaited.", None),
        ("        No. of columns   how many attribute columns the layer has.", None),
        ("", None),
        ("2.  \"Columns\"  — one row per column, for every layer. Use this to find out what is IN a layer.", bold),
        ("        Theme, Layer     which layer the column belongs to (match these to the Layers sheet).", None),
        ("        Column           the column name in the database.", None),
        ("        Type             the data type (text, number, date, geometry, …).", None),
        ("        Description/unit what the column means and, for measurements, its unit.", None),
        ("", None),
        ("How to find something", head),
        ("Looking for a topic? Filter the Layers sheet by Theme, or use Find (Ctrl+F) on the Description.", None),
        ("Found a layer and want its contents? Filter the Columns sheet by that Layer name.", None),
        ("", None),
        ("Theme legend", head),
    ]
    for text, font in rows:
        info.append([text])
        if font is not None:
            info.cell(row=info.max_row, column=1).font = font
    # theme legend table
    info.append(["Code", "Theme", "Layers"])
    for cell in info[info.max_row]:
        cell.font = head_font
        cell.fill = head_fill
    for t in sorted({r["theme"] for r in records}):
        info.append([t, THEME_LABELS.get(t, t),
                     sum(1 for r in records if r["theme"] == t)])
        info.cell(row=info.max_row, column=1).font = bold
    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 60
    info.column_dimensions["C"].width = 10


_URL_RE = re.compile(r"https?://[^\s<]+")


def _linkify(escaped: str) -> str:
    """Wrap bare URLs in an already HTML-escaped string in anchor tags."""
    return _URL_RE.sub(lambda m: f'<a href="{m.group(0)}" target="_blank" rel="noopener">{m.group(0)}</a>', escaped)


def _section_html(sections: dict[str, str]) -> str:
    out = []
    for head in SECTION_HEADERS:
        if head not in sections or not sections[head].strip():
            continue
        items = parse_bullets(sections[head])
        if not items:
            continue
        lis = "".join(f"<li>{_linkify(html.escape(it))}</li>" for it in items)
        block = f"<ul>{lis}</ul>"
        out.append(f'<div class="sec"><span class="sechead">{html.escape(head)}</span>{block}</div>')
    return "".join(out)


def render_html(records: list[dict[str, Any]], built: str, path: Path) -> None:
    themes = sorted({r["theme"] for r in records})
    n_layers = len(records)

    sections_html = []
    for t in themes:
        colour = THEME_COLOURS.get(t, BRAND_INK)
        fg = text_on(colour)
        label = THEME_LABELS.get(t, t)
        blurb = THEME_DESC.get(t, "")
        layers = [r for r in records if r["theme"] == t]
        cards = []
        for r in layers:
            title = first_sentence(r["description"])
            col_rows = "".join(
                f"<tr><td><code>{html.escape(c['name'])}</code></td>"
                f"<td><code>{html.escape(c['type'])}</code></td>"
                f"<td>{html.escape(c['comment'] or '')}</td></tr>"
                for c in r["columns"]
            )
            # Dashboard not yet released — no deep-link for now.
            map_link = ""
            cards.append(f"""
        <details class="layer" data-theme="{t}" data-search="{html.escape((r['table'] + ' ' + r['description']).lower())}" style="border-left:4px solid {colour}">
          <summary><span class="badge" style="background:{colour};color:{fg}">{t}</span>
            <code>{html.escape(r['table'])}</code><span class="title"> — {html.escape(title)}</span></summary>
          <p class="desc">{html.escape(r['description'])}</p>
          {map_link}
          {_section_html(r['sections'])}
          <table class="cols">
            <thead><tr><th>Column</th><th>Type</th><th>Description / unit</th></tr></thead>
            <tbody>{col_rows}</tbody>
          </table>
        </details>""")
        sections_html.append(f"""
  <section class="theme-group" data-theme="{t}" id="theme-{t}">
    <h2 style="background:{colour};color:{fg}"><span class="tcode">{t}</span> {html.escape(label)} <span class="tn">{len(layers)}</span></h2>
    <p class="blurb">{html.escape(blurb)}</p>
    {''.join(cards)}
  </section>""")

    theme_opts = "".join(
        f'<option value="{t}">{t} · {html.escape(THEME_LABELS.get(t, t))} ({sum(1 for r in records if r["theme"] == t)})</option>'
        for t in themes
    )

    page = f"""<!DOCTYPE html>
<html lang="en-GB"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>uk_baseline data catalogue</title>
<style>
 :root {{ --ink:{BRAND_INK}; --accent:{BRAND_ACCENT}; --line:#d9d3c4; --bg:{BRAND_BG}; }}
 * {{ box-sizing:border-box; }}
 body {{ font:15px/1.55 -apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
        color:{BRAND_BLACK}; margin:0; background:var(--bg); }}
 header {{ background:var(--ink); color:#fff; padding:26px 32px; border-bottom:5px solid var(--accent); }}
 header h1 {{ margin:0 0 4px; font-size:23px; letter-spacing:.2px; }}
 header p {{ margin:0; opacity:.8; font-size:13px; }}
 .controls {{ position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line);
              padding:14px 32px; display:flex; gap:12px; flex-wrap:wrap; z-index:5; }}
 .controls input, .controls select {{ padding:8px 10px; border:1px solid var(--line);
              border-radius:6px; font-size:14px; background:#fff; }}
 .controls input {{ flex:1; min-width:240px; }}
 main {{ padding:8px 32px 60px; max-width:1120px; }}
 .count {{ color:#6b6357; font-size:13px; margin:14px 0 4px; }}
 .theme-group {{ margin:26px 0 8px; }}
 .theme-group > h2 {{ font-size:15px; margin:0 0 2px; padding:9px 14px; border-radius:8px;
             display:flex; align-items:center; gap:10px; }}
 .theme-group .tcode {{ font-weight:800; letter-spacing:1px; opacity:.85; font-size:13px; }}
 .theme-group .tn {{ margin-left:auto; font-size:12px; opacity:.85;
             background:rgba(255,255,255,.25); padding:1px 9px; border-radius:10px; }}
 .blurb {{ color:#6b6357; font-size:12.5px; margin:4px 6px 12px; }}
 details.layer {{ background:#fff; border:1px solid var(--line); border-radius:8px;
                  margin:0 0 9px; }}
 details.layer summary {{ cursor:pointer; padding:13px 16px; font-size:14.5px;
                  list-style:none; display:flex; align-items:center; gap:10px; flex-wrap:wrap; }}
 details.layer summary::-webkit-details-marker {{ display:none; }}
 details[open] summary {{ border-bottom:1px solid var(--line); }}
 .title {{ color:#5a5346; font-weight:400; }}
 .badge {{ font-size:11px; font-weight:800; padding:2px 9px; border-radius:10px; letter-spacing:.5px; }}
 .desc {{ padding:0 16px; color:#33302a; }}
 .maplink {{ padding:2px 16px; margin:2px 0; }}
 .maplink a {{ color:#005dff; font-weight:600; text-decoration:none; }}
 .sec {{ padding:2px 16px; }}
 .sechead {{ display:inline-block; font-size:11px; font-weight:800; letter-spacing:.5px;
             color:var(--accent); margin-top:8px; }}
 .sec ul {{ margin:4px 0 6px; padding-left:20px; }}
 .sec a {{ color:#005dff; word-break:break-all; }}
 table.cols {{ width:calc(100% - 32px); margin:10px 16px 16px; border-collapse:collapse; font-size:13px; }}
 table.cols th, table.cols td {{ text-align:left; padding:6px 8px; border-bottom:1px solid var(--line);
             vertical-align:top; }}
 table.cols th {{ background:var(--bg); }}
 code {{ background:var(--bg); padding:1px 5px; border-radius:4px; font-size:12.5px; }}
 mark {{ background:#dadf50; }}
</style></head>
<body>
<header>
  <h1>uk_baseline data catalogue</h1>
  <p>{n_layers} layers across {len(themes)} themes &middot; generated from the database metadata on {built}</p>
</header>
<div class="controls">
  <input id="q" type="search" placeholder="Search layers by name or description…" autocomplete="off">
  <select id="theme"><option value="">All themes</option>{theme_opts}</select>
</div>
<main>
  <div class="count" id="count"></div>
  {''.join(sections_html)}
</main>
<script>
 const layers = Array.from(document.querySelectorAll('details.layer'));
 const groups = Array.from(document.querySelectorAll('.theme-group'));
 const q = document.getElementById('q'), themeSel = document.getElementById('theme'),
       count = document.getElementById('count');
 function apply() {{
   const term = q.value.trim().toLowerCase(), th = themeSel.value;
   let shown = 0;
   for (const el of layers) {{
     const okTheme = !th || el.dataset.theme === th;
     const okTerm = !term || el.dataset.search.includes(term);
     const show = okTheme && okTerm;
     el.style.display = show ? '' : 'none';
     if (show) shown++;
   }}
   for (const g of groups) {{
     const anyVisible = Array.from(g.querySelectorAll('details.layer'))
       .some(d => d.style.display !== 'none');
     g.style.display = anyVisible ? '' : 'none';
   }}
   count.textContent = shown + ' of {n_layers} layers shown';
 }}
 q.addEventListener('input', apply); themeSel.addEventListener('change', apply); apply();
</script>
</body></html>"""
    path.write_text(page, encoding="utf-8")
    log.info("Wrote %s", path.name)


def _section_md(sections: dict[str, str]) -> str:
    out = []
    for head in SECTION_HEADERS:
        if head not in sections or not sections[head].strip():
            continue
        out.append(f"**{head}**\n")
        for bullet in parse_bullets(sections[head]):
            out.append(f"- {bullet}")
        out.append("")
    return "\n".join(out)


def render_mkdocs(records: list[dict[str, Any]], built: str) -> None:
    DOCS.mkdir(parents=True, exist_ok=True)
    for stale in DOCS.rglob("*.md"):        # drop pages for removed/renamed layers
        try:
            stale.unlink()
        except OSError as exc:
            log.warning("Could not remove stale page %s: %s", stale, exc)

    # Show the FULL layer name in the left nav — table names are long and
    # underscore-joined, which MkDocs Material truncates with an ellipsis.
    # Allow them to wrap (break on underscores) and never clip.
    css_dir = DOCS / "stylesheets"
    css_dir.mkdir(exist_ok=True)
    (css_dir / "extra.css").write_text(
        "/* Catalogue nav: show full, wrapped layer names (no ellipsis truncation). */\n"
        ".md-nav__item .md-ellipsis {\n"
        "  white-space: normal;\n"
        "  overflow: visible;\n"
        "  overflow-wrap: anywhere;\n"
        "  word-break: break-word;\n"
        "  line-height: 1.3;\n"
        "}\n"
        ".md-nav__link { align-items: flex-start; }\n"
        "/* Layer short name: sits between the page title and the table name. */\n"
        ".layer-short {\n"
        "  font-size: 1.15rem;\n"
        "  font-weight: 600;\n"
        "  color: #e9511d;\n"
        "  margin: -0.4rem 0 0.2rem;\n"
        "}\n"
        "/* Brand-orange admonition (note) instead of Material blue. */\n"
        ".md-typeset .admonition.note,\n"
        ".md-typeset details.note { border-color: #e9511d; }\n"
        ".md-typeset .note > .admonition-title { background-color: #fbe6db; }\n"
        ".md-typeset .note > .admonition-title::before { background-color: #e9511d; }\n"
        "/* Homepage footer block (data info, contact, version). */\n"
        ".cat-footer { display: flex; flex-wrap: wrap; gap: 2rem; margin-top: 3rem;\n"
        "  padding-top: 1.5rem; border-top: 1px solid #d9d3c4; color: #6b6357; font-size: .8rem; }\n"
        ".cat-footer > div { flex: 1; min-width: 200px; }\n"
        ".cat-footer strong { color: #2b242c; }\n"
        ".cat-footer a { color: #e9511d; }\n",
        encoding="utf-8",
    )
    themes = sorted({r["theme"] for r in records})

    # Landing page
    rows = "\n".join(
        f"| **{t}** | [{THEME_LABELS.get(t, t)}]({t}/index.md) | {THEME_DESC.get(t, '')} | {sum(1 for r in records if r['theme'] == t)} |"
        for t in themes
    )
    (DOCS / "index.md").write_text(
        f"""# uk_baseline data catalogue

A guide to every layer in Prior + Partners' `uk_baseline` spatial database.

## Getting started

**What is `uk_baseline`?** A single, central store of the UK-wide spatial datasets the
practice uses across projects — administrative boundaries, demographics, environment,
economy, transport, heritage and more. Everyone works from the same maintained source,
instead of copies scattered across folders and machines.

**How to use this catalogue.** Browse by theme below, or use the search box. Every layer
has its own page with a description, a short name, its database table name, a preview map,
where the data came from (source, licence and documentation), and a table of every column
with its meaning and units.

**New to the database?** The **[Staff Data Management Handbook](PP_Staff_Data_Management_Handbook.pdf){{ target="_blank" }}** explains what the geodatabase
is, how to connect QGIS, and how to use the data on a project — it opens straight in your browser.

**{len(records)} layers across {len(themes)} themes.**

| Code | Theme | Typical contents | Layers |
|---|---|---|---:|
{rows}

!!! note "How to read a layer page"
    Each layer page shows a short description, where the data came from
    (source, documentation links, licence), the area and detail it covers, and a
    table of every column with its meaning and units — all taken verbatim from
    the database documentation.

<div class="cat-footer">
  <div><strong>About this catalogue</strong><br>
  Built from the <code>uk_baseline</code> PostGIS database — one page per layer, drawn from each layer's own documentation.</div>
  <div><strong>Contact</strong><br>
  Digital Innovation team, Prior + Partners<br>
  <a href="mailto:info@priorpartners.com">info@priorpartners.com</a></div>
  <div><strong>Version</strong><br>
  Last updated {built}<br>
  Developed by the Digital Innovation team</div>
</div>
""",
        encoding="utf-8",
    )

    nav_themes = []
    for t in themes:
        tdir = DOCS / t
        tdir.mkdir(exist_ok=True)
        layers = [r for r in records if r["theme"] == t]

        def _title(rec: dict[str, Any]) -> str:
            return first_sentence(rec["description"]).rstrip(".").strip() or rec["table"]

        lst = "\n".join(f"- [{_title(r)}]({r['table']}.md) · `{r['table']}`"
                        for r in layers)
        blurb = THEME_DESC.get(t, "")
        (tdir / "index.md").write_text(
            f"# {THEME_LABELS.get(t, t)} ({t})\n\n*{blurb}*\n\n{len(layers)} layer{'s' if len(layers) != 1 else ''}.\n\n{lst}\n",
            encoding="utf-8",
        )
        # nav label = plain-English title (table name stays on the page)
        nav_pages = [{_title(r): f"{t}/{r['table']}.md"} for r in layers]
        nav_themes.append({f"{THEME_LABELS.get(t, t)} ({t})":
                           [f"{t}/index.md"] + nav_pages})

        for r in layers:
            col_rows = "\n".join(
                f"| `{c['name']}` | `{c['type']}` | {(c['comment'] or '').replace('|', '\\|')} |"
                for c in r["columns"]
            )
            fs = first_sentence(r["description"])
            rest = r["description"][len(fs):].strip()
            rest_block = f"{rest}\n\n" if rest else ""
            # Static styling snapshot (PNG) shows only where built — themes are
            # converted one at a time. The Dashboard deep-link is offered for
            # every layer (it serves them all). With use_directory_urls the layer
            # page is at /<THEME>/<table>/, so the image (site /maps/) is two
            # levels up; mkdocs does not rewrite raw-HTML attributes.
            img = ""
            if has_map(r["table"]):
                img = (f'<img src="../../maps/{r["table"]}.png" '
                       f'alt="Styling preview of {r["table"]}" loading="lazy" '
                       f'style="width:100%;border:1px solid #d9d3c4;'
                       f'border-radius:8px;margin:6px 0 4px;">\n\n')
            # Dashboard not yet released — show the preview image only (no deep-link).
            map_block = f"{img}" if img else ""
            sn = r.get("short_name", "")
            short_html = f'<p class="layer-short">{html.escape(sn)}</p>\n\n' if sn else ""
            (tdir / f"{r['table']}.md").write_text(
                f"""# {_title(r)}

{short_html}`{r['table']}`

{map_block}{rest_block}{_section_md(r['sections'])}

## Columns

| Column | Type | Description / unit |
|---|---|---|
{col_rows}
""",
                encoding="utf-8",
            )

    nav = [{"Home": "index.md"},
           {"📕 Staff Handbook (PDF)":
            "https://prior-partners.github.io/PP_uk_baseline/PP_Staff_Data_Management_Handbook.pdf"}] + nav_themes
    mkdocs_yml = OUT / "mkdocs.yml"
    yaml_lines = [
        "site_name: uk_baseline data catalogue",
        "site_description: A guide to every layer in the uk_baseline database.",
        f"copyright: Developed by the Digital Innovation team, Prior + Partners — last updated {built}",
        "theme:",
        "  name: material",
        "  palette:",
        "    primary: deep orange",
        "    accent: deep orange",
        "  features:",
        "    - navigation.instant",
        "    - navigation.tracking",
        "    - toc.integrate",
        "    - search.highlight",
        "    - search.suggest",
        "    - content.code.copy",
        "plugins:",
        "  - search",
        "markdown_extensions:",
        "  - admonition",
        "  - attr_list",
        "  - tables",
        "  - toc:",
        "      permalink: true",
        "extra_css:",
        "  - stylesheets/extra.css",
        "nav:",
    ]
    yaml_lines += _yaml_nav(nav, indent=1)
    mkdocs_yml.write_text("\n".join(yaml_lines) + "\n", encoding="utf-8")
    log.info("Wrote MkDocs source under %s and mkdocs.yml", DOCS)


def _yaml_nav(items: list[Any], indent: int) -> list[str]:
    pad = "  " * indent
    lines: list[str] = []
    for item in items:
        if isinstance(item, str):                 # bare path, no title
            lines.append(f"{pad}- {item}")
            continue
        for key, val in item.items():
            k = '"' + key.replace('\\', '\\\\').replace('"', '\\"') + '"'  # quote: titles may contain : & ,
            if isinstance(val, str):
                lines.append(f'{pad}- {k}: {val}')
            else:
                lines.append(f'{pad}- {k}:')
                lines += _yaml_nav(val, indent + 1)
    return lines


# --- main ------------------------------------------------------------------
def main() -> None:
    built = datetime.now(timezone.utc).strftime("%d %B %Y")
    records = read_metadata()
    shorts = read_short_names()
    for r in records:
        r["short_name"] = shorts.get(r["table"], "")
    log.info("Read %d layers from %s (%d short names from register)",
             len(records), SCHEMA, sum(1 for r in records if r["short_name"]))

    OUT.mkdir(parents=True, exist_ok=True)
    SHARE.mkdir(parents=True, exist_ok=True)

    def _guarded(fn, path: Path, label: str) -> None:
        """Run a renderer; if the target is locked (open in Excel), warn and skip."""
        try:
            fn(records, built, path)
        except PermissionError:
            log.warning("SKIPPED %s — %s is open/locked. Close it and re-run to refresh.",
                        label, path.name)

    # Single-file deliverables on the practice share, for colleagues.
    _guarded(render_xlsx, SHARE / "uk_baseline_catalogue.xlsx", "Excel")
    _guarded(render_html, SHARE / "uk_baseline_catalogue.html", "HTML")
    # MkDocs site source stays local (built / published separately).
    render_mkdocs(records, built)

    # Remove any earlier local copies of the single-file outputs.
    for stale in (OUT / "uk_baseline_catalogue.xlsx", OUT / "uk_baseline_catalogue.html"):
        stale.unlink(missing_ok=True)

    fp = fingerprint(records)
    (OUT / ".catalogue_fingerprint.json").write_text(
        json.dumps({"fingerprint": fp, "built": built,
                    "n_layers": len(records), "schema": SCHEMA}, indent=2),
        encoding="utf-8",
    )
    log.info("Fingerprint %s", fp[:16])
    log.info("Single-file deliverables: %s", SHARE)
    log.info("MkDocs site source: %s", OUT)


if __name__ == "__main__":
    main()
